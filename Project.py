from openpyxl import load_workbook
import re

INPUT_FILE = "Project Excel.xlsx"
SHEET_NAME = "Sheet1"

COLUMN_ALIASES = {
    "item_name": ["item", "item name", "product name", "name"],
    "sku": ["sku", "item code", "product code"],
    "current_stock": ["stock", "current stock",  "quantity"],
    "reorder_level": ["reorder level", "minimum stock", "min stock"],
    "target_stock": ["target stock",  "desired stock", "max stock"],
    "avg_daily_sales": ["avg daily sales", "daily sales", "sales per day"],
    "lead_time_days": ["lead time", "lead time days", "supplier lead time"],
    "unit_cost": ["unit cost", "cost", "item cost", "purchase cost"],
    "supplier": ["supplier", "vendor", "supplier name"],
}

REQUIRED_COLUMNS = ["item_name", "current_stock", "reorder_level", "target_stock"]
OUTPUT_SHEET_NAME = "Priority Results"

def normalize_header(text): # this cleans up the header text for better matching
    text = str(text).strip().lower()
    text = re.sub(r"[_-]+", " ", text)
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def load_sheet(file_path, sheet_name): # this loads the Excel file and returns the workbook and the specified sheet
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    return workbook, sheet


def read_headers(sheet, header_row=1): # this reads the header row and returns a dictionary mapping column index to header text
    headers = {}
    for col in range(1, sheet.max_column + 1):
        headers[col] = sheet.cell(row=header_row, column=col).value
    return headers


def map_columns(headers, column_aliases):# this maps the headers to the internal column names based on the provided aliases and returns a dictionary of internal name to column index
    mapped_columns = {}

    normalized_headers = {}
    for col_index, header in headers.items():
        if header is not None:
            normalized_headers[col_index] = normalize_header(header) # this cleans up the header text for better matching

    for internal_name, alias_list in column_aliases.items():
        normalized_aliases = [normalize_header(alias) for alias in alias_list]

        for col_index, header_text in normalized_headers.items():
            if header_text in normalized_aliases:
                mapped_columns[internal_name] = col_index
                break

    return mapped_columns

# After cleaning up the text for use we need a error check to make sure all the required columns are present in the mapped columns. If any are missing, we raise a ValueError with a message indicating which columns are missing.

def validate_required_columns(mapped_columns, required_columns):
    missing = []
    for col in required_columns:
        if col not in mapped_columns:
            missing.append(col)

    if missing:
        raise ValueError(f"Missing required columns: {missing}")

def to_number(value, default=0): # this function attempts to convert a value to a float, returning a default value if the conversion fails or if the value is None or an empty string
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default
    

def read_inventory_rows(sheet, mapped_columns, header_row=1): # this function reads the inventory data rows from the sheet based on the mapped columns and returns a list of item dictionaries containing the relevant data for each item
    items = []

    for row in range(header_row + 1, sheet.max_row + 1):
        item_name_col = mapped_columns["item_name"]
        item_name = sheet.cell(row=row, column=item_name_col).value

        # Skip completely empty rows
        if item_name is None or str(item_name).strip() == "":
            continue

        item = {
            "excel_row": row,
            "item_name": item_name,
            "sku": sheet.cell(row=row, column=mapped_columns["sku"]).value if "sku" in mapped_columns else "",
            "supplier": sheet.cell(row=row, column=mapped_columns["supplier"]).value if "supplier" in mapped_columns else "",
            "current_stock": to_number(sheet.cell(row=row, column=mapped_columns["current_stock"]).value),
            "reorder_level": to_number(sheet.cell(row=row, column=mapped_columns["reorder_level"]).value),
            "target_stock": to_number(sheet.cell(row=row, column=mapped_columns["target_stock"]).value),
            "avg_daily_sales": to_number(
                sheet.cell(row=row, column=mapped_columns["avg_daily_sales"]).value
            ) if "avg_daily_sales" in mapped_columns else 0,
            "lead_time_days": to_number(
                sheet.cell(row=row, column=mapped_columns["lead_time_days"]).value
            ) if "lead_time_days" in mapped_columns else 0,
            "unit_cost": to_number(
                sheet.cell(row=row, column=mapped_columns["unit_cost"]).value
            ) if "unit_cost" in mapped_columns else 0,
        }

        items.append(item)

    return items


def compute_priority_metrics(items):# this function computes various priority metrics for each item based on the inventory data and returns a list of enriched item dictionaries with the computed metrics and a priority score for ranking
    ranked_items = []

    for item in items:
        current_stock = item["current_stock"]
        reorder_level = item["reorder_level"]
        target_stock = item["target_stock"]
        avg_daily_sales = item["avg_daily_sales"]
        lead_time_days = item["lead_time_days"]
        unit_cost = item["unit_cost"]

        shortage_to_target = max(0, target_stock - current_stock)
        below_reorder = current_stock < reorder_level
        reorder_gap = max(0, reorder_level - current_stock)

        days_of_stock_left = None
        if avg_daily_sales > 0:
            days_of_stock_left = current_stock / avg_daily_sales

        lead_time_demand = avg_daily_sales * lead_time_days

        # Basic scoring model
        priority_score = 0

        if below_reorder:
            priority_score += 50

        priority_score += reorder_gap * 5
        priority_score += shortage_to_target * 2
        priority_score += lead_time_demand * 3

        if days_of_stock_left is not None and days_of_stock_left < lead_time_days:
            priority_score += 25

        estimated_reorder_cost = shortage_to_target * unit_cost

        enriched = item.copy()
        enriched.update({
            "shortage_to_target": shortage_to_target,
            "reorder_gap": reorder_gap,
            "below_reorder": below_reorder,
            "days_of_stock_left": round(days_of_stock_left, 2) if days_of_stock_left is not None else "",
            "lead_time_demand": round(lead_time_demand, 2),
            "estimated_reorder_qty": shortage_to_target,
            "estimated_reorder_cost": round(estimated_reorder_cost, 2),
            "priority_score": round(priority_score, 2),
        })

        ranked_items.append(enriched)

    ranked_items.sort(
        key=lambda x: (
            x["priority_score"],
            x["estimated_reorder_qty"]
        ),
        reverse=True
    )

    for rank, item in enumerate(ranked_items, start=1):
        item["priority_rank"] = rank

    return ranked_items

def write_priority_results(workbook, ranked_items, output_sheet_name): # this function writes the ranked items with their computed metrics and priority scores back to a new sheet in the workbook, creating headers and populating the rows with the relevant data for each item
    if output_sheet_name in workbook.sheetnames:
        del workbook[output_sheet_name]

    result_sheet = workbook.create_sheet(output_sheet_name)

    output_headers = [
        "Priority Rank",
        "Item Name",
        "SKU",
        "Supplier",
        "Current Stock",
        "Reorder Level",
        "Target Stock",
        "Avg Daily Sales",
        "Lead Time Days",
        "Days of Stock Left",
        "Lead Time Demand",
        "Reorder Gap",
        "Estimated Reorder Qty",
        "Unit Cost",
        "Estimated Reorder Cost",
        "Below Reorder",
        "Priority Score",
        "Excel Row",
    ]

    for col, header in enumerate(output_headers, start=1):
        result_sheet.cell(row=1, column=col, value=header)

    for row_index, item in enumerate(ranked_items, start=2):
        values = [
            item["priority_rank"],
            item["item_name"],
            item["sku"],
            item["supplier"],
            item["current_stock"],
            item["reorder_level"],
            item["target_stock"],
            item["avg_daily_sales"],
            item["lead_time_days"],
            item["days_of_stock_left"],
            item["lead_time_demand"],
            item["reorder_gap"],
            item["estimated_reorder_qty"],
            item["unit_cost"],
            item["estimated_reorder_cost"],
            item["below_reorder"],
            item["priority_score"],
            item["excel_row"],
        ]

        for col_index, value in enumerate(values, start=1):
            result_sheet.cell(row=row_index, column=col_index, value=value)






    
def main():
    workbook, sheet = load_sheet(INPUT_FILE, SHEET_NAME)

    print(f"Loaded workbook: {INPUT_FILE}")
    print(f"Using sheet: {SHEET_NAME}")
    print("-" * 50)

    headers = read_headers(sheet)
    print("Raw headers:")
    for col_index, header in headers.items():
        print(f"  Column {col_index}: {header}")

    print("-" * 50)

    mapped_columns = map_columns(headers, COLUMN_ALIASES)
    print("Mapped columns:")
    for internal_name, col_index in mapped_columns.items():
        print(f"  {internal_name} -> Column {col_index}")

    print("-" * 50)

    validate_required_columns(mapped_columns, REQUIRED_COLUMNS)
    print("All required columns found.")

    print("-" * 50)

    items = read_inventory_rows(sheet, mapped_columns)
    print(f"Read {len(items)} inventory rows.")

    print("\nSample parsed rows:")
    for item in items[:5]:
        print(item)

    print("-" * 50)

    ranked_items = compute_priority_metrics(items)

    print("Top priority items:")
    for item in ranked_items:
        print(
            f"Rank {item['priority_rank']}: "
            f"{item['item_name']} | "
            f"Score={item['priority_score']} | "
            f"Current={item['current_stock']} | "
            f"Reorder={item['reorder_level']} | "
            f"Target={item['target_stock']} | "
            f"Reorder Qty={item['estimated_reorder_qty']}"
        )

    print("-" * 50)

    write_priority_results(workbook, ranked_items, OUTPUT_SHEET_NAME)
    workbook.save(INPUT_FILE)

    print(f"Results written to sheet: {OUTPUT_SHEET_NAME}")
    print("Test completed successfully.")

if __name__ == "__main__":
    main()